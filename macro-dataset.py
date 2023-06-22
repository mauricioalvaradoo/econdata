import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from econdata import BCRP, YFinance

pd.options.display.max_columns = None



# Data ========================================================================
# PBI por el lado del gasto
data_1 = BCRP.get_data(
    {
         'PN02518AQ': 'Consumo privado (var. % anual)',
         'PN02522AQ': 'Inversión privada (var. % anual)',
         'PN02519AQ': 'Consumo público (var. % anual)',
         'PN02523AQ': 'Inversión pública (var. % anual)',
         'PN02524AQ': 'Exportaciones (var. % anual)',
         'PN02525AQ': 'Importaciones (var. % anual)',
         'PN02517AQ': 'Demanda Interna (var. % anual)',
         'PN02527AQ': 'Demanda Interna sin inventarios (var. % anual)',
         'PN02526AQ': 'PBI (var. % anual)',
    },
    '1980Q1',
    '2022Q4'
)

# PBI por sectores
data_2 = BCRP.get_data(
    {
         'PN02499AQ': 'PBI Agropecuario (var. % anual)',
         'PN02500AQ': 'PBI Pesca (var. % anual)',
         'PN02501AQ': 'PBI Minería e Hidrocarburos (var. % anual)',
         'PN02502AQ': 'PBI Manufactura (var. % anual)',
         'PN02503AQ': 'PBI Electricidad y Agua (var. % anual)',
         'PN02504AQ': 'PBI Construcción (var. % anual)',
         'PN02505AQ': 'PBI Comercio (var. % anual)',
         'PN02506AQ': 'PBI Servicios (var. % anual)',
         'PN02507AQ': 'PBI (var. % anual)',
         'PN37680AQ': 'PBI Primario (var. % anual)',
         'PN37681AQ': 'PBI No Primario (var. % anual)',
    },
    '1980Q1',
    '2022Q4'
)

# Inflación 
data_3 = BCRP.get_data(
    {
         'PN01277PM': 'Inflación Sin Alimentos y Energía (%)',
         'PN09819PM': 'Inflación Alimentos y Energía (%)',
         'PN01279PM': 'Inflación Subyacente (%)',
         'PN09820PM': 'Inflación No Subyacente (%)',
         'PN01281PM': 'Inflación Transables (%)',
         'PN01283PM': 'Inflación No Transables (%)',
         'PN01285PM': 'Inflación No Transables Sin Alimentos (%)',
         'PN09822PM': 'Inflación Alimentos y Bebidas (%)',
         'PN09823PM': 'Inflación Sin Alimentos y Bebidas (%)',
         'PN09824PM': 'Inflación Subyacente Sin Alimentos y Bebidas (%)',
         'PN09821PM': 'Inflación Importada (%)',
         'PN01287PM': 'Inflación al por Mayor (%)',
         'PN01273PM': 'Inflación (%)',
    },
    '1993-01',
    '2022-12'
)

# Empleo y Salarios
data_4 = BCRP.get_data(
    {
         'PN38052GM': 'PEA Ocupada - Por Edad - 14 a 24 años (Miles de personas, promedio móvil tres meses)',
         'PN38053GM': 'PEA Ocupada - Por Edad - 25 a 44 años (Miles de personas, promedio móvil tres meses)',
         'PN38054GM': 'PEA Ocupada - Por Edad - 45 a más años (Miles de personas, promedio móvil tres meses)',
         'PN38055GM': 'PEA Ocupada - Por Categoría Ocupacional - Independiente (Miles de personas, promedio móvil tres meses)',
         'PN38056GM': 'PEA Ocupada - Por Categoría Ocupacional - Dependiente (Miles de personas, promedio móvil tres meses)',
         'PN38057GM': 'PEA Ocupada - Por Categoría Ocupacional - Trabajador no Remunerado (Miles de personas, promedio móvil tres meses)',
         'PN38058GM': 'PEA Ocupada - Por Tamaño de Empresa - De 1 a 10 trabajadores (Miles de personas, promedio móvil tres meses)',
         'PN38059GM': 'PEA Ocupada - Por Tamaño de Empresa - De 11 a 50 trabajadores (Miles de personas, promedio móvil tres meses)',
         'PN38060GM': 'PEA Ocupada - Por Tamaño de Empresa - De 51 y más (Miles de personas, promedio móvil tres meses)',
         'PN38061GM': 'PEA Adecuadamente Empleada (Miles de personas, promedio móvil tres meses)',
         'PN38062GM': 'PEA Subempleada (Miles de personas, promedio móvil tres meses)',
         'PN38069GM': 'Coeficiente de Ocupación (%, promedio móvil tres meses)',
         'PN38064GM': 'Tasa de Desempleo (%) - Por Género - Hombre (%, promedio móvil tres meses)',
         'PN38065GM': 'Tasa de Desempleo (%) - Por Género - Mujer (%, promedio móvil tres meses)',
         'PN38066GM': 'Tasa de Desempleo (%) - Por Grupos de Edad - 14 a 24 años (%, promedio móvil tres meses)',
         'PN38067GM': 'Tasa de Desempleo (%) - Por Grupos de Edad - 25 a 44 años (%, promedio móvil tres meses)',
         'PN38068GM': 'Tasa de Desempleo (%) - Por Grupos de Edad - 45 a más años (%, promedio móvil tres meses)',
         'PN38070GM': 'Ingreso Mensual (Promedio móvil tres meses)',
         'PN02124PM': 'Remuneración Mínima Vital (S/)',
         'PN37696PM': 'Ingreso promedio del sector formal privado (S/)',
         'PN31885GM': 'Masa salarial del sector formal total (S/ Millones)',
    },
    '2001-01',
    '2022-12'
)

# Coyuntura
data_5 = BCRP.get_data(
    {
        'PD38041AM':' Índice de Venta Respecto al Mes Anterior (base=50)',
        'PD38042AM': 'Índice de Inventarios Respecto al Mes Anterior (base=50)',
        'PD38043AM': 'Índice de Órdenes de Compra Respecto al Mes Anterior (base=50)',
        'PD38044AM': 'Índice de la Situación Actual del Negocio (base=50)',
        'PD38045AM': 'Índice de Expectativas de la Economía a 3 Meses (base=50)',
        'PD38046AM': 'Índice de Expectativas del Sector a 3 Meses (base=50)',
        'PD38047AM': 'Índice de Expectativas de la Demanda a 3 Meses (base=50)',
        'PD37981AM': 'Índice de Expectativas de la Economía a 12 Meses (base=50)',
        'PD39751AM': 'Índice de Expectativas de Demanda por sus Productos a 12 Meses (base=50)',
        'PD39752AM': 'Índice de Expectativas del Sector a 12 Meses (base=50)',
        'PD37966AM': 'Consumo Interno de Electricidad (GWH)',
        'PD38040AM': 'Consumo Interno de Electricidad sin Minería (GWH)',
    },
    '2001-01',
    '2022-12'
)

# Monetaria
data_6 = BCRP.get_data(
    {
         'PD04722MM': 'Tasa de Referencia de Política Monetaria (%)',
         'PN07839NM': 'Tasa de Interés Interbancaria (%)',
         'PN00493MM': 'Tasa de Encaje',
         'PN00489MM': 'Emisión primaria por el BCRP (S/ Millones)',
         'PN00480MM': 'Circulante (S/ Millones)',
         'PN00494MM': 'Multiplicador Monetario (%)',
         'PN06481IM': 'Reservas Internacionales Netas (US$ Millones)',
         'PN01234PM': 'Tipo de cambio U.S.',
         'PN01235PM': 'Tipo de cambio Euro',
         'PN01240PM': 'Tipo de cambio Yuan',
    },
    '1998-01',
    '2022-12'
)

# Fiscal
data_7 = BCRP.get_data(
    {
        'PN02318FM': 'Ingresos del Gobierno Central - Impuesto a las Importaciones (S/ Millones)',
        'PN02319FM': 'Ingresos del Gobierno Central - Impuesto General a las Ventas (S/ Millones)',
        'PN02322FM': 'Ingresos del Gobierno Central - Impuesto Selectivo al Consumo (S/ Millones)',
        'PN02325FM': 'Ingresos del Gobierno Central - Otros Ingresos (S/ Millones)',
        'PN02327FM': 'Ingresos del Gobierno Central - Ingresos No Tributarios (S/ Millones)',
        'PN02223FM': 'Gastos del Gobierno Central - Remuneraciones (S/ Millones)',
        'PN02224FM': 'Gastos del Gobierno Central - Bienes y Servicios (S/ Millones)',
        'PN02225FM': 'Gastos del Gobierno Central - Transferencias (S/ Millones)',
        'PN02232FM': 'Gastos del Gobierno Central - Formación Bruta de Capital (S/ Millones)',
        'PN02236FM': 'Gastos del Gobierno Central - Intereses de la Deuda Interna (S/ Millones)',
        'PN02237FM': 'Gastos del Gobierno Central - Intereses de la Deuda Externa (S/ Millones)',
    },
    '2004-01',
    '2022-12'
)

# Balanza Comercial
data_8 = BCRP.get_data(
    {
        'PN38739BM': 'Exportaciones de Productos Tradicionales - Pesqueros (US$ Millones, FOB)',
        'PN38740BM': 'Exportaciones de Productos Tradicionales - Agrícolas (US$ Millones, FOB)',
        'PN38741BM': 'Exportaciones de Productos Tradicionales - Mineros (US$ Millones, FOB)',
        'PN38742BM': 'Exportaciones de Productos Tradicionales - Petróleo y Gas Natural (US$ Millones, FOB)',
        'PN38744BM': 'Exportaciones de Productos no Tradicionales - Agropecuarios (US$ Millones, FOB)',
        'PN38745BM': 'Exportaciones de Productos no Tradicionales - Pesqueros (US$ Millones, FOB)',
        'PN38746BM': 'Exportaciones de Productos no Tradicionales - Textiles (US$ Millones, FOB)',
        'PN38747BM': 'Exportaciones de Productos no Tradicionales - Maderas y Papeles, y sus Manufacturas (US$ Millones, FOB)',
        'PN38748BM': 'Exportaciones de Productos no Tradicionales - Químicos (US$ Millones, FOB)',
        'PN38749BM': 'Exportaciones de Productos no Tradicionales - Minerales no Metálicos (US$ Millones, FOB)',
        'PN38750BM': 'Exportaciones de Productos no Tradicionales - Sidero - Metalúrgicos y Joyería (US$ Millones, FOB)',
        'PN38751BM': 'Exportaciones de Productos no Tradicionales - Metal - Mecánicos (US$ Millones, FOB)',
        'PN38891BM': 'Importaciones de Bienes de Consumo (US$ Millones, FOB)',
        'PN38894BM': 'Importaciones de Insumos (US$ Millones, FOB)',
        'PN38898BM': 'Importaciones de Bienes de Capital (US$ Millones, FOB)',
        'PN38915BM': 'Índice de Precios Nominales - Exportaciones (Índice 2007=100)',
        'PN38919BM': 'Índice de Precios Nominales - Importaciones (Índice 2007=100)',
    },
    '2013-01',
    '2022-12'
)

# Commodities
data_9 = BCRP.get_data(
    {
         'PN01652XM': 'Cobre LME',
         'PN01654XM': 'Oro LME',
         'PN01655XM': 'Plata H. Harman',
         'PN01657XM': 'Zinc LME',
         'PN01653XM': 'Estaño LME',
         'PN01656XM': 'Plomo LME',
         'PN01660XM': 'Petróleo WTI',
         'PN01661XM': 'Trigo EEUU',
         'PN01662XM': 'Maíz EEUU',
         'PN01664XM': 'Aceite Soya EEUU',
         'PN01649XM': 'Harina de Pescado Hamburgo',
    },
    '1994-01',
    '2022-12'
)

# Stocks
data_10_1 = BCRP.get_data(
    {
         'PN01142MM': 'Índice General BVL (base 31/12/91 = 100)',
         'PN01143MM': 'Índice Selectivo BVL (base 31/12/91 = 100)',
         
    },
    '1998-01',
    '2022-12'
)

data_10_2 = YFinance.get_data(
    {   
        '^GSPC': 'S&P 500',
        '^DJI': 'Dow Jones Industrial Average',
        '^IXIC': 'NASDAQ Composite',
        '^FTSE': 'FTSE 100',
        '^N225': 'Nikkei 225',
    },
    fechaini='1998-01-01',
    fechafin='2022-12-31'
)

data_10_2.index = pd.to_datetime(data_10_2.index)
data_10_2 = data_10_2.resample('M', axis = 0).mean()
data_10_2.index = data_10_2.index.to_period('M').to_timestamp()
data_10 = pd.concat([data_10_1, data_10_2], axis=1)



# Guardado ====================================================================
with pd.ExcelWriter('macro-dataset.xlsx', engine='openpyxl') as excel_writer:
    data_1.to_excel(excel_writer, sheet_name='PBI_Gasto')
    data_2.to_excel(excel_writer, sheet_name='PBI_Sectores')
    data_3.to_excel(excel_writer, sheet_name='Inflación')
    data_4.to_excel(excel_writer, sheet_name='Empleo_Salarios')
    data_5.to_excel(excel_writer, sheet_name='Coyuntura')
    data_6.to_excel(excel_writer, sheet_name='Monetario')
    data_7.to_excel(excel_writer, sheet_name='Fiscal')
    data_8.to_excel(excel_writer, sheet_name='Balanza_Comercial')
    data_9.to_excel(excel_writer, sheet_name='Commodities')
    data_10.to_excel(excel_writer, sheet_name='Stocks')

    for sheet_name in excel_writer.sheets:
        sheet = excel_writer.sheets[sheet_name]
        sheet.row_dimensions[1].height = 120  # Tamaño fila 1
        for cell in sheet[1]:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True) # Ajustar texto
        
        sheet.freeze_panes = 'B2' # Inmobilizar paneles

        column_a = sheet['A']
        column_a_number = openpyxl.utils.column_index_from_string(column_a[0].column_letter)
        column_a_width = 15
        sheet.column_dimensions[get_column_letter(column_a_number)].width = column_a_width
        for cell in column_a:
            cell.number_format = 'dd/mm/yyyy' # Formato de la columna A: fecha

        for column in sheet.columns:
            column_number = openpyxl.utils.column_index_from_string(column[0].column_letter)
            column_width = 15
            sheet.column_dimensions[get_column_letter(column_number)].width = column_width
            
            for cell in column:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0' # Un decimal
